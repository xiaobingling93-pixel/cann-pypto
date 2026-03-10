#!/usr/bin/env python3
# Copyright (c) Huawei Technologies Co., Ltd. 2024-2026. All rights reserved.

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import time
import urllib.request
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any, TypedDict

SKILL_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_PATH = SKILL_ROOT / "references" / "rule_ch.xlsx"
DEFAULT_OFFICIAL_URL = (
    "https://raw.gitcode.com/cann/infrastructure/blobs/"
    "572341d2f1ed4522ba9225020fc00e77f734f09c/docs/SC/CodeArts-Check/rule_ch.xlsx"
)


RULE_ID_LOOSE_RE = re.compile(r"\b(G\.[A-Z]{3}\.\d+)\b", re.IGNORECASE)


class Args(argparse.Namespace):
    rule_ids: list[str] = []
    rules_file: Path | None = None
    stdin: bool = False
    from_violations_json: str | None = None
    category: str | None = None
    list: bool = False
    language: str = "python"
    excel: Path = DEFAULT_EXCEL_PATH
    download: bool = False
    download_url: str = DEFAULT_OFFICIAL_URL
    download_retries: int = 3
    format: str = "markdown"


class RuleRow(TypedDict, total=False):
    language: str
    id: str
    title: str
    name: str
    level: object
    suggestion: object
    correct_example: object
    wrong_example: object


def eprint(msg: str) -> None:
    logging.error(msg)


def normalize_rule_id(value: str) -> str:
    raw = (value or "").strip().upper()
    if not raw:
        return ""

    m = RULE_ID_LOOSE_RE.search(raw)
    if m:
        raw = m.group(1).upper()
    else:
        token = raw.split()[0]
        if re.fullmatch(r"[A-Z]{3}\.\d+", token):
            raw = f"G.{token}"
        elif re.fullmatch(r"G\.[A-Z]{3}\.\d+", token):
            raw = token
        else:
            return ""

    parts = raw.split(".")
    if len(parts) != 3:
        return ""
    num = parts[2]
    if num.isdigit() and len(num) == 1:
        parts[2] = f"0{num}"
    return ".".join(parts)


def parse_rule_cell(cell_value: Any) -> tuple[str, str]:
    text = "" if cell_value is None else str(cell_value)
    text = text.replace("\r", " ").replace("\n", " ").strip()
    rid = normalize_rule_id(text)
    if not rid:
        return "", text
    title = RULE_ID_LOOSE_RE.sub("", text, count=1).strip()
    title = re.sub(r"^[\s:：\-]+", "", title).strip()
    return rid, title


def read_rules_file(path: Path) -> list[str]:
    rule_ids: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        rid = normalize_rule_id(s)
        if rid:
            rule_ids.append(rid)
    return rule_ids


def download_excel(url: str, output_path: Path, retries: int) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            eprint(f"Downloading ({attempt}/{retries}): {url}")
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = resp.read()
            output_path.write_bytes(data)
            eprint(f"Saved: {output_path}")
            return
        except Exception as e:
            last_error = e
            if attempt < retries:
                wait_s = 2 ** (attempt - 1)
                eprint(f"Download failed: {e}; retry in {wait_s}s")
                time.sleep(wait_s)
    raise RuntimeError(f"Download failed after {retries} attempts: {last_error}")


def infer_cols(headers: Sequence[object]) -> dict[str, int]:
    hdr = [("" if cell is None else str(cell)).strip().lower() for cell in headers]

    def find_any(keys: Iterable[str]) -> int | None:
        for i, h in enumerate(hdr):
            if not h:
                continue
            for k in keys:
                if k in h:
                    return i
        return None

    cols: dict[str, int] = {}
    cols["name"] = find_any(["规则名称", "rule name", "rule", "规则"]) or 0
    level = find_any(["问题级别", "级别", "severity", "level"])
    correct = find_any(["正确示例", "正例", "correct"])
    wrong = find_any(["错误示例", "反例", "wrong", "incorrect"])
    suggestion = find_any(["修改建议", "建议", "suggestion", "fix"])
    if level is not None:
        cols["level"] = level
    if correct is not None:
        cols["correct_example"] = correct
    if wrong is not None:
        cols["wrong_example"] = wrong
    if suggestion is not None:
        cols["suggestion"] = suggestion
    return cols


def load_rules(excel_path: Path, languages: list[str]) -> list[RuleRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl") from e

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    out: list[RuleRow] = []
    for lang in languages:
        if lang not in wb.sheetnames:
            continue
        ws = wb[lang]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            continue

        cols = infer_cols(header)
        for row in rows_iter:
            if not row:
                continue
            name_idx = cols["name"]
            cell = row[name_idx] if name_idx < len(row) else None
            if not cell:
                continue
            rid, title = parse_rule_cell(cell)
            if not rid:
                continue
            item: RuleRow = {
                "language": lang,
                "id": rid,
                "title": title,
                "name": str(cell).strip(),
            }

            for k in ["level", "suggestion", "correct_example", "wrong_example"]:
                idx = cols.get(k)
                if idx is not None and idx < len(row):
                    v = row[idx]
                    if v is not None and str(v).strip() != "":
                        item[k] = v
            out.append(item)
    return out


def dedupe(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in items:
        if x in seen:
            continue
        seen.add(x)
        out.append(x)
    return out


def extract_rule_ids_from_violations_json(obj: object) -> list[str]:
    if isinstance(obj, dict) and isinstance(obj.get("grouped"), dict):
        out: list[str] = []
        for key in obj["grouped"].keys():
            rid = normalize_rule_id(str(key))
            if rid:
                out.append(rid)
        return out

    if isinstance(obj, dict) and isinstance(obj.get("violations"), list):
        obj = obj["violations"]
    elif isinstance(obj, dict) and isinstance(obj.get("issues"), list):
        obj = obj["issues"]

    if not isinstance(obj, list):
        return []

    out: list[str] = []
    for v in obj:
        if not isinstance(v, dict):
            continue
        rid = normalize_rule_id(
            str(
                v.get("rule_id")
                or v.get("ruleId")
                or v.get("rule")
                or v.get("rule_code")
                or ""
            )
        )
        if not rid and isinstance(v.get("detail"), dict):
            detail = v["detail"]
            rid = normalize_rule_id(str(detail.get("rule") or detail.get("rule_id") or ""))
        if not rid and isinstance(v.get("message"), str):
            rid = normalize_rule_id(v["message"])
        if rid:
            out.append(rid)
    return out


def index_by_id(rows: list[RuleRow]) -> dict[str, list[RuleRow]]:
    idx: dict[str, list[RuleRow]] = {}
    for r in rows:
        rule_id = r.get("id")
        if isinstance(rule_id, str) and rule_id:
            idx.setdefault(rule_id, []).append(r)
    return idx


def format_markdown(by_id: dict[str, list[RuleRow]], source_url: str) -> str:
    lines: list[str] = []
    lines.append("# CodeCheck 规则查询结果")
    lines.append("")
    lines.append(f"> 官方来源: {source_url}")
    lines.append("")
    for rid, items in by_id.items():
        lines.append(f"## {rid}")
        if not items:
            lines.append("未找到该规则。")
            lines.append("")
            lines.append("---")
            lines.append("")
            continue
        for r in items:
            lines.append(f"**语言**: {r.get('language', '')}")
            title = r.get("title")
            if title:
                lines.append(f"**标题**: {title}")
            level = r.get("level")
            if level is not None:
                lines.append(f"**级别**: {level}")

            suggestion = r.get("suggestion")
            if suggestion:
                lines.append("")
                lines.append("**修改建议**:")
                lines.append(str(suggestion).strip())

            correct_example = r.get("correct_example")
            if correct_example:
                lines.append("")
                lines.append("**正确示例**:")
                lines.append("```")
                lines.append(str(correct_example).rstrip())
                lines.append("```")

            wrong_example = r.get("wrong_example")
            if wrong_example:
                lines.append("")
                lines.append("**错误示例**:")
                lines.append("```")
                lines.append(str(wrong_example).rstrip())
                lines.append("```")
            lines.append("")
        lines.append("---")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    parser = argparse.ArgumentParser(description="CodeCheck 规则查询工具（支持批量）")
    parser.add_argument("rule_ids", nargs="*", help="规则 ID（可多个），如 G.CLS.06 G.FMT.02")
    parser.add_argument("--rules-file", type=Path, help="从文件读取规则 ID（每行一个，支持 # 注释）")
    parser.add_argument("--stdin", action="store_true", help="从 stdin 读取规则 ID（每行一个）")
    parser.add_argument(
        "--rules",
        metavar="RULES",
        help="Comma-separated rule IDs or path to violations JSON (e.g., G.CLS.06,G.LOG.02 or violations.json)",
    )
    parser.add_argument(
        "--from-violations-json",
        dest="from_violations_json",
        metavar="PATH",
        help="从 violations JSON 提取规则 ID（传文件路径或 '-' 表示 stdin）",
    )
    parser.add_argument("--category", "-c", help="按类别查询，如 CLS/FMT/ERR")
    parser.add_argument("--list", action="store_true", help="列出规则（包含标题）")
    parser.add_argument(
        "--language",
        "-l",
        choices=["python", "c++", "c", "all"],
        default="python",
        help="语言/工作表（默认 python；all 表示所有）",
    )
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH, help="Excel 文件路径")
    parser.add_argument("--download", action="store_true", help="强制下载/更新官方 Excel")
    parser.add_argument("--download-url", default=DEFAULT_OFFICIAL_URL, help="官方下载 URL")
    parser.add_argument("--download-retries", type=int, default=3, help="下载重试次数")
    parser.add_argument("--format", "-f", choices=["json", "markdown"], default="markdown", help="输出格式")

    args = parser.parse_args()

    excel_path: Path = args.excel
    if args.download or not excel_path.exists():
        try:
            download_excel(args.download_url, excel_path, retries=max(1, int(args.download_retries)))
        except Exception as e:
            eprint(f"Error: {e}")
            return 2

    languages = ["python", "c++", "c"] if args.language == "all" else [args.language]
    try:
        rows = load_rules(excel_path, languages)
    except Exception as e:
        eprint(f"Error: {e}")
        return 2

    if args.list:
        items = sorted(rows, key=lambda r: (r.get("language", ""), r.get("id", "")))
        if args.format == "json":
            logging.info(
                json.dumps(
                    {
                        "source": args.download_url,
                        "excel": str(excel_path),
                        "total": len(items),
                        "rules": items,
                    },
                    ensure_ascii=False,
                    indent=2,
                )
            )
        else:
            by_lang: dict[str, list[str]] = {}
            for r in items:
                lang = str(r.get("language", ""))
                rid = str(r.get("id", ""))
                by_lang.setdefault(lang, []).append(f"{rid} {r.get('title','')}")
            logging.info(f"# CodeCheck 规则列表\n\n> 官方来源: {args.download_url}\n")
            for lang, xs in by_lang.items():
                logging.info(f"## {lang}\n")
                for x in xs:
                    logging.info(f"- {x}".rstrip())
                logging.info("")
        eprint(f"Found {len(items)} rules")
        return 0

    if args.category:
        cat = (args.category or "").strip().upper()
        items = [r for r in rows if r.get("id", "").startswith(f"G.{cat}.")]
        items.sort(key=lambda r: (r.get("id", ""), r.get("language", "")))
        if args.format == "json":
            logging.info(
                json.dumps(
                    {
                        "source": args.download_url,
                        "excel": str(excel_path),
                        "query": {"category": cat, "language": args.language},
                        "total": len(items),
                        "rules": items,
                    },
                    ensure_ascii=False,
                    indent=2,
                )
            )
        else:
            by_id: dict[str, list[RuleRow]] = {}
            for r in items:
                rule_id = str(r.get("id", ""))
                if rule_id:
                    by_id.setdefault(rule_id, []).append(r)
            logging.info(format_markdown(by_id, args.download_url))
        eprint(f"Found {len(items)} rules")
        return 0

    rule_ids: list[str] = []
    for x in args.rule_ids:
        rid = normalize_rule_id(x)
        if rid:
            rule_ids.append(rid)

    if args.rules_file:
        try:
            rule_ids.extend(read_rules_file(args.rules_file))
        except Exception as e:
            eprint(f"Error: {e}")
            return 2

    if args.stdin:
        for line in sys.stdin.read().splitlines():
            rid = normalize_rule_id(line)
            if rid:
                rule_ids.append(rid)

    if args.rules:
        rules_arg = (args.rules or "").strip()
        if rules_arg:
            rules_path = Path(rules_arg)
            if rules_path.exists() and rules_path.is_file() and rules_path.suffix.lower() == ".json":
                try:
                    obj = json.loads(rules_path.read_text(encoding="utf-8"))
                    rule_ids.extend(extract_rule_ids_from_violations_json(obj))
                except Exception as e:
                    eprint(f"Error: {e}")
                    return 2
            else:
                for part in rules_arg.split(","):
                    rid = normalize_rule_id(part)
                    if rid:
                        rule_ids.append(rid)

    if args.from_violations_json:
        src = args.from_violations_json
        try:
            raw = sys.stdin.read() if src == "-" else Path(src).read_text(encoding="utf-8")
            obj = json.loads(raw)
            rule_ids.extend(extract_rule_ids_from_violations_json(obj))
        except Exception as e:
            eprint(f"Error: {e}")
            return 2

    rule_ids = dedupe([x for x in rule_ids if x])
    if not rule_ids:
        parser.print_help()
        return 1

    idx = index_by_id(rows)
    by_id: dict[str, list[RuleRow]] = {rid: idx.get(rid, []) for rid in rule_ids}

    if args.format == "json":
        logging.info(
            json.dumps(
                {
                    "source": args.download_url,
                    "excel": str(excel_path),
                    "query": {"rule_ids": rule_ids, "language": args.language},
                    "total": sum(len(v) for v in by_id.values()),
                    "by_id": by_id,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
    else:
        logging.info(format_markdown(by_id, args.download_url))

    missing = [rid for rid, items in by_id.items() if not items]
    eprint(f"Queried {len(rule_ids)} rule ids, missing {len(missing)}")
    if missing:
        eprint("Missing: " + ", ".join(missing[:30]) + (" ..." if len(missing) > 30 else ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
